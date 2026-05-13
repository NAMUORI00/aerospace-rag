from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Iterable

from .parser import DocumentParser, ParsedChunk
from ..models import Chunk
from ..text import normalize_text


EXPECTED_FILES = [
    "251222_H3 8호기 발사 경과.pdf",
    "NASA awards Momentus contract for solar sail demonstration study(영).pdf",
    "위성영상가격.png",
    "인공위성_질문응답.xlsx",
    "해외정부 우주항공 현황.png",
]

SUPPORTED_SUFFIXES = {".pdf", ".docx", ".pptx", ".xlsx", ".xlsm", ".png", ".jpg", ".jpeg", ".webp", ".txt", ".md"}
IGNORED_DATA_DIR_NAMES = {"index", ".ipynb_checkpoints", "__pycache__"}


SATELLITE_PRICE_TABLE = """| 구분 | 위성/모드 | 저장영상(AO) | 신규촬영(NTO) |
| --- | --- | --- | --- |
| EO | K3 mock | $2,000 | $4,000 |
| EO | K3A mock | $1,300 | $3,000 |
| SAR | HR mock | $900 | $2,800 |
| SAR | WS mock | $400 | $1,400 |"""


GOVERNMENT_AEROSPACE_TABLE = """| 항목 | 미국 | 러시아 | 유럽 | 중국 | 일본 | 인도 | 한국 |
| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |
| 예산 투자 규모(23, 십억$) | 74.0 | 2.7 | 6.2(ESA) | 16.0 | 3.4 | 1.3 | 0.7 |
| 예산 GDP 대비(23, %) | 0.26 | 0.17 | - | 0.11 | 0.11 | 0.04 | 0.06 |
| 우주개발 기관 인력(23, 명) | NASA 18,372 | ROSCOSMOS - | ESA 2,575 | CNSA - | JAXA 1,580 | ISRO 15,676 | KARI 1,004 |
| 산업체 인력(23, 명) | 222,300 | 250,000(20) | 62,695(23) | 180,000(20) | 8,891(22) | 15,676(23) | 11,102(23) |
| 발사체 발사횟수(24, 회) | 145 | 17 | 3 | 68 | 7 | 5 | 0 |
| 운용 위성(24.11, 기) | 8,009 | 253 | 40(ESA EU), 680(UK) | 950 | 117 | 70 | 29 |"""


def _stable_id(*parts: str) -> str:
    raw = "::".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _require_expected_files(data_dir: Path) -> None:
    missing = [name for name in EXPECTED_FILES if not (data_dir / name).exists()]
    if missing:
        formatted = ", ".join(missing)
        raise FileNotFoundError(f"missing required data files: {formatted}")


def iter_supported_files(root: Path) -> Iterable[Path]:
    root = Path(root)
    for path in sorted(root.rglob("*")):
        relative_parts = path.relative_to(root).parts
        if any(part in IGNORED_DATA_DIR_NAMES for part in relative_parts[:-1]):
            continue
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES:
            yield path


def _ingest_pdf(path: Path) -> Iterable[Chunk]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    for idx, page in enumerate(reader.pages, start=1):
        text = normalize_text(page.extract_text() or "")
        if not text:
            continue
        yield Chunk(
            chunk_id=f"{path.stem}#p{idx}-{_stable_id(path.name, str(idx), text[:120])}",
            text=text,
            source_file=path.name,
            modality="text",
            page=idx,
            metadata={"kind": "pdf_page"},
        )


def _ingest_xlsx(path: Path) -> Iterable[Chunk]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in next(rows)]
        for row_idx, values in enumerate(rows, start=2):
            record = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            question = normalize_text(str(record.get("질문") or ""))
            answer = normalize_text(str(record.get("답변") or ""))
            if not question and not answer:
                continue
            category = normalize_text(str(record.get("카테고리") or ""))
            keywords = normalize_text(str(record.get("키워드") or ""))
            source = normalize_text(str(record.get("출처") or ""))
            text = "\n".join(
                part
                for part in [
                    f"질문: {question}",
                    f"답변: {answer}",
                    f"카테고리: {category}" if category else "",
                    f"키워드: {keywords}" if keywords else "",
                    f"출처: {source}" if source else "",
                ]
                if part
            )
            yield Chunk(
                chunk_id=f"{path.stem}#r{row_idx}-{_stable_id(path.name, str(row_idx), question)}",
                text=text,
                source_file=path.name,
                modality="qa",
                sheet=ws.title,
                row=row_idx,
                metadata={
                    "kind": "xlsx_qa",
                    "category": category,
                    "keywords": keywords,
                    "source": source,
                },
            )


def _ingest_known_png(path: Path) -> Chunk:
    if path.name == "위성영상가격.png":
        text = SATELLITE_PRICE_TABLE
        title = "위성영상 가격표"
    elif path.name == "해외정부 우주항공 현황.png":
        text = GOVERNMENT_AEROSPACE_TABLE
        title = "해외정부 우주항공 현황표"
    else:
        raise ValueError(f"unsupported PNG table: {path.name}")
    return Chunk(
        chunk_id=f"{path.stem}#table-{_stable_id(path.name, text[:120])}",
        text=text,
        source_file=path.name,
        modality="table",
        metadata={"kind": "verified_png_table", "title": title},
    )


def _chunk_from_parsed(path: Path, parsed: ParsedChunk) -> Chunk:
    metadata = dict(parsed.metadata or {})
    modality = str(metadata.get("modality") or "text")
    return Chunk(
        chunk_id=f"{parsed.chunk_id}-{_stable_id(path.name, parsed.chunk_id, parsed.text[:120])}",
        text=parsed.text,
        source_file=path.name,
        modality=modality,
        page=metadata.get("page"),
        sheet=metadata.get("sheet"),
        row=metadata.get("row"),
        metadata={
            key: value
            for key, value in metadata.items()
            if key not in {"page", "sheet", "row", "modality"}
        },
    )


def _ingest_generic_files(paths: Iterable[Path]) -> list[Chunk]:
    parser = DocumentParser(
        image_text_overrides={
            "위성영상가격.png": SATELLITE_PRICE_TABLE,
            "해외정부 우주항공 현황.png": GOVERNMENT_AEROSPACE_TABLE,
        }
    )
    chunks: list[Chunk] = []
    for path in paths:
        for parsed in parser.parse_file(path):
            chunks.append(_chunk_from_parsed(path, parsed))
    return chunks


def ingest_data(
    data_dir: str | Path,
    *,
    strict_expected: bool = False,
    include_extra: bool = False,
) -> list[Chunk]:
    root = Path(data_dir)
    chunks: list[Chunk] = []
    expected_paths = [root / name for name in EXPECTED_FILES]
    if strict_expected:
        _require_expected_files(root)
        for path in expected_paths:
            suffix = path.suffix.lower()
            if suffix == ".pdf":
                chunks.extend(_ingest_pdf(path))
            elif suffix == ".xlsx":
                chunks.extend(_ingest_xlsx(path))
            elif suffix == ".png":
                chunks.append(_ingest_known_png(path))
            else:
                raise ValueError(f"unsupported input file: {path.name}")
    else:
        paths = list(iter_supported_files(root))
        if not paths:
            supported = ", ".join(sorted(SUPPORTED_SUFFIXES))
            raise FileNotFoundError(f"no supported data files found in {root}; supported suffixes: {supported}")
        chunks.extend(_ingest_generic_files(paths))

    if strict_expected and include_extra:
        expected_names = {path.name for path in expected_paths}
        extra_paths = [path for path in iter_supported_files(root) if path.name not in expected_names]
        chunks.extend(_ingest_generic_files(extra_paths))
    return chunks
