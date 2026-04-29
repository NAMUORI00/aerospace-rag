from __future__ import annotations

import base64
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .text import normalize_text


_TOKEN_RE = re.compile(r"\S+")


@dataclass(frozen=True)
class ParsedChunk:
    chunk_id: str
    text: str
    metadata: dict[str, Any]


class DocumentParser:
    """Dependency-light document parser modeled after SmartFarm ingest.

    It prefers optional `unstructured` when available, but the default path is
    pure Python and works for this repo's Colab/local deployment contract.
    """

    def __init__(self, *, image_text_overrides: dict[str, str] | None = None) -> None:
        self.chunk_token_size = 1200
        self.chunk_token_overlap = 100
        self.strategy = "hi_res"
        self.extract_images = True
        self.infer_table_structure = True
        self.languages = ["kor", "eng"]
        self.image_text_overrides = dict(image_text_overrides or {})
        self._converter = self._build_unstructured()

    def _build_unstructured(self):
        try:
            from unstructured.partition.auto import partition
        except Exception:
            return None
        return partition

    def _chunk_text(
        self,
        *,
        doc_stem: str,
        text: str,
        source_doc: str,
        modality: str = "text",
        asset_ref: str | None = None,
        chunk_index_seed: int = 0,
        page: int | None = None,
        sheet: str | None = None,
        row: int | None = None,
        table_html_ref: str | None = None,
        formula_latex_ref: str | None = None,
        image_b64_ref: str | None = None,
    ) -> list[ParsedChunk]:
        raw = normalize_text(str(text or ""))
        if not raw:
            return []
        base_meta: dict[str, Any] = {
            "source_doc": source_doc,
            "modality": modality,
            "asset_ref": asset_ref,
            "page": page,
            "sheet": sheet,
            "row": row,
            "table_html_ref": table_html_ref,
            "formula_latex_ref": formula_latex_ref,
            "image_b64_ref": image_b64_ref,
        }

        if modality in {"table", "image", "formula", "qa"}:
            return [
                ParsedChunk(
                    chunk_id=f"{doc_stem}#{modality[:1]}{chunk_index_seed}",
                    text=raw,
                    metadata=base_meta,
                )
            ]

        tokens = _TOKEN_RE.findall(raw)
        if not tokens:
            return []
        token_size = int(self.chunk_token_size)
        overlap = min(int(self.chunk_token_overlap), max(0, token_size - 1))
        step = max(1, token_size - overlap)
        chunks: list[ParsedChunk] = []
        idx = chunk_index_seed
        for start in range(0, len(tokens), step):
            window = tokens[start : start + token_size]
            if not window:
                break
            chunks.append(
                ParsedChunk(
                    chunk_id=f"{doc_stem}#t{idx}",
                    text=" ".join(window),
                    metadata=base_meta,
                )
            )
            idx += 1
            if start + token_size >= len(tokens):
                break
        return chunks

    def _detect_modality(self, category: str) -> str:
        c = str(category or "").lower()
        if "table" in c:
            return "table"
        if "image" in c or "figure" in c or "picture" in c:
            return "image"
        if "formula" in c or "equation" in c or "math" in c:
            return "formula"
        return "text"

    def _safe_attr(self, obj, name: str, default=None):
        try:
            return getattr(obj, name, default)
        except Exception:
            return default

    def _to_image_b64(self, value) -> str | None:
        if value is None:
            return None
        if isinstance(value, str) and value.strip():
            return value.strip()
        if isinstance(value, (bytes, bytearray)):
            return base64.b64encode(bytes(value)).decode("utf-8")
        return None

    def _parse_with_unstructured(self, path: Path) -> list[ParsedChunk]:
        if self._converter is None:
            return []
        try:
            elements = self._converter(
                filename=str(path),
                strategy=self.strategy,
                extract_images_in_pdf=bool(self.extract_images),
                infer_table_structure=bool(self.infer_table_structure),
                languages=self.languages,
            )
        except Exception:
            return []

        chunks: list[ParsedChunk] = []
        seed = 0
        for idx, element in enumerate(elements or []):
            text = str(self._safe_attr(element, "text", "") or "").strip()
            category = str(self._safe_attr(element, "category", "") or "")
            meta = self._safe_attr(element, "metadata", None)
            modality = self._detect_modality(category)
            page_no = None
            table_html = None
            formula_latex = None
            image_b64 = None
            if meta is not None:
                page_no = self._safe_attr(meta, "page_number", None)
                table_html = str(self._safe_attr(meta, "text_as_html", "") or "").strip() or None
                formula_latex = str(self._safe_attr(meta, "text_as_latex", "") or "").strip() or None
                image_b64 = self._to_image_b64(self._safe_attr(meta, "image_base64", None))
            if not text:
                text = table_html or formula_latex or ""
            if not text:
                continue
            parsed = self._chunk_text(
                doc_stem=path.stem,
                text=text,
                source_doc=path.name,
                modality=modality,
                asset_ref=f"page:{page_no}#idx:{idx}" if page_no else f"asset:{idx}",
                chunk_index_seed=seed,
                page=int(page_no) if page_no else None,
                table_html_ref=table_html,
                formula_latex_ref=formula_latex,
                image_b64_ref=image_b64,
            )
            chunks.extend(parsed)
            seed += len(parsed)
        return chunks

    def _parse_pdf(self, path: Path) -> list[ParsedChunk]:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        chunks: list[ParsedChunk] = []
        for idx, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            chunks.extend(
                self._chunk_text(
                    doc_stem=path.stem,
                    text=text,
                    source_doc=path.name,
                    modality="text",
                    chunk_index_seed=idx,
                    page=idx,
                )
            )
        return chunks

    def _parse_xlsx(self, path: Path) -> list[ParsedChunk]:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        chunks: list[ParsedChunk] = []
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [str(v or "").strip() for v in next(rows)]
            except StopIteration:
                continue
            for row_idx, values in enumerate(rows, start=2):
                record = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                question = normalize_text(str(record.get("질문") or record.get("question") or ""))
                answer = normalize_text(str(record.get("답변") or record.get("answer") or ""))
                if question or answer:
                    text = "\n".join(part for part in [f"질문: {question}", f"답변: {answer}"] if part.strip())
                    modality = "qa"
                else:
                    cells = [f"{headers[i]}: {values[i]}" for i in range(min(len(headers), len(values))) if values[i] is not None]
                    text = "\n".join(cells)
                    modality = "table"
                chunks.extend(
                    self._chunk_text(
                        doc_stem=path.stem,
                        text=text,
                        source_doc=path.name,
                        modality=modality,
                        chunk_index_seed=row_idx,
                        sheet=ws.title,
                        row=row_idx,
                    )
                )
        return chunks

    def _parse_image(self, path: Path) -> list[ParsedChunk]:
        text = self.image_text_overrides.get(path.name, "")
        if not text:
            text = f"이미지 파일: {path.name}"
        return self._chunk_text(
            doc_stem=path.stem,
            text=text,
            source_doc=path.name,
            modality="table" if path.name in self.image_text_overrides else "image",
            asset_ref=str(path),
        )

    def _parse_text(self, path: Path) -> list[ParsedChunk]:
        text = path.read_text(encoding="utf-8", errors="ignore")
        return self._chunk_text(doc_stem=path.stem, text=text, source_doc=path.name, modality="text")

    def parse_file(self, path: str | Path) -> list[ParsedChunk]:
        p = Path(path)
        if p.suffix.lower() in {".pdf", ".docx", ".pptx"}:
            parsed = self._parse_with_unstructured(p)
            if parsed:
                return parsed
        suffix = p.suffix.lower()
        if suffix == ".pdf":
            return self._parse_pdf(p)
        if suffix in {".xlsx", ".xlsm"}:
            return self._parse_xlsx(p)
        if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            return self._parse_image(p)
        if suffix in {".txt", ".md"}:
            return self._parse_text(p)
        raise ValueError(f"unsupported input file: {p.name}")

