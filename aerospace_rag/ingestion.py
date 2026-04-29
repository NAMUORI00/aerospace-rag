from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Iterable

from .models import Chunk
from .text import normalize_text


EXPECTED_FILES = [
    "251222_H3 8호기 발사 경과.pdf",
    "NASA awards Momentus contract for solar sail demonstration study(영).pdf",
    "위성영상가격.png",
    "인공위성_질문응답.xlsx",
    "해외정부 우주항공 현황.png",
]


SATELLITE_PRICE_TABLE = """| 구분 | 위성/모드 | 해상도(m) | 저장영상(AO) | 신규촬영(NTO) |
| --- | --- | ---: | --- | --- |
| EO | K2 | 1 | 15 x 15 scene, $900, 1,260,000원 | 신규촬영 없음 |
| EO | K3 | 0.7 | 16 x 16 scene, $2,048, 2,867,200원, km^2 $8 | 16 x 16 scene, $4,096, 5,734,400원, km^2 $16 |
| EO | K3A | 0.55 | 13 x 13 scene, $1,352, 1,892,800원, km^2 $8 | 13 x 13 scene, $3,042, 4,258,800원, km^2 $18 |
| SAR | HR(UH) | 1(0.85) | 5 x 5 scene, $900, 1,260,000원 | 5 x 5 scene, $2,800, 3,920,000원 |
| SAR | ST(ES) | 3(2.5) | 30 x 30 scene, $600, 840,000원 | 30 x 30 scene, $1,800, 2,520,000원 |
| SAR | WS | 20 | 100 x 100 scene, $400, 560,000원 | 100 x 100 scene, $1,400, 1,960,000원 |

메모: 환율 1400원 기준. tile 판매 시 주문 최소크기 10 x 10. 스테레오 촬영은 scene을 2장 붙인 것으로 가격도 2배."""


GOVERNMENT_AEROSPACE_TABLE = """| 항목 | 미국 | 러시아 | 유럽 | 중국 | 일본 | 인도 | 한국 |
| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |
| 예산 투자 규모(23, 십억$) | 74.0 | 2.7 | 6.2(ESA) | 16.0 | 3.4 | 1.3 | 0.7 |
| 예산 GDP 대비(23, %) | 0.26 | 0.17 | - | 0.11 | 0.11 | 0.04 | 0.06 |
| 우주개발 기관 인력(23, 명) | NASA 18,372 | ROSCOSMOS - | ESA 2,575 | CNSA - | JAXA 1,580 | ISRO 15,676 | KARI 1,004 |
| 산업체 인력(23, 명) | 222,300 | 250,000(20) | 62,695(23) | 180,000(20) | 8,891(22) | 15,676(23) | 11,102(23) |
| 발사체 발사횟수(24, 회) | 145 | 17 | 3 | 68 | 7 | 5 | 0 |
| 운용 위성(24.11, 기) | 8,009 | 253 | 40(ESA EU), 680(UK) | 950 | 117 | 70 | 29 |"""


def _stable_id(*parts: str) -> str:
    raw = "::".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _require_expected_files(data_dir: Path) -> None:
    missing = [name for name in EXPECTED_FILES if not (data_dir / name).exists()]
    if missing:
        formatted = ", ".join(missing)
        raise FileNotFoundError(f"missing required data files: {formatted}")


def _ingest_pdf(path: Path) -> Iterable[Chunk]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    for idx, page in enumerate(reader.pages, start=1):
        text = normalize_text(page.extract_text() or "")
        if not text:
            continue
        yield Chunk(
            chunk_id=f"{path.stem}#p{idx}-{_stable_id(path.name, str(idx), text[:120])}",
            text=text,
            source_file=path.name,
            modality="text",
            page=idx,
            metadata={"kind": "pdf_page"},
        )


def _ingest_xlsx(path: Path) -> Iterable[Chunk]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        headers = [str(v or "").strip() for v in next(rows)]
        for row_idx, values in enumerate(rows, start=2):
            record = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            question = normalize_text(str(record.get("질문") or ""))
            answer = normalize_text(str(record.get("답변") or ""))
            if not question and not answer:
                continue
            category = normalize_text(str(record.get("카테고리") or ""))
            keywords = normalize_text(str(record.get("키워드") or ""))
            source = normalize_text(str(record.get("출처") or ""))
            text = "\n".join(
                part
                for part in [
                    f"질문: {question}",
                    f"답변: {answer}",
                    f"카테고리: {category}" if category else "",
                    f"키워드: {keywords}" if keywords else "",
                    f"출처: {source}" if source else "",
                ]
                if part
            )
            yield Chunk(
                chunk_id=f"{path.stem}#r{row_idx}-{_stable_id(path.name, str(row_idx), question)}",
                text=text,
                source_file=path.name,
                modality="qa",
                sheet=ws.title,
                row=row_idx,
                metadata={
                    "kind": "xlsx_qa",
                    "category": category,
                    "keywords": keywords,
                    "source": source,
                },
            )


def _ingest_known_png(path: Path) -> Chunk:
    if path.name == "위성영상가격.png":
        text = SATELLITE_PRICE_TABLE
        title = "위성영상 가격표"
    elif path.name == "해외정부 우주항공 현황.png":
        text = GOVERNMENT_AEROSPACE_TABLE
        title = "해외정부 우주항공 현황표"
    else:
        raise ValueError(f"unsupported PNG table: {path.name}")
    return Chunk(
        chunk_id=f"{path.stem}#table-{_stable_id(path.name, text[:120])}",
        text=text,
        source_file=path.name,
        modality="table",
        metadata={"kind": "verified_png_table", "title": title},
    )


def ingest_data(data_dir: str | Path) -> list[Chunk]:
    root = Path(data_dir)
    _require_expected_files(root)
    chunks: list[Chunk] = []
    for name in EXPECTED_FILES:
        path = root / name
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            chunks.extend(_ingest_pdf(path))
        elif suffix == ".xlsx":
            chunks.extend(_ingest_xlsx(path))
        elif suffix == ".png":
            chunks.append(_ingest_known_png(path))
        else:
            raise ValueError(f"unsupported input file: {path.name}")
    return chunks
